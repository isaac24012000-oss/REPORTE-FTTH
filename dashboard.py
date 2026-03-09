import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np
import os

st.set_page_config(
    page_title="Reporte Bitel FTTH",
    page_icon="📞",
    layout="wide",
    initial_sidebar_state="expanded"
)
# Actualizado 02/03/2026 - Preparado para MARZO 2026

# ============= CARGA DE DATOS DEL EXCEL =============

@st.cache_data(ttl=3600)
def load_mantra_data():
    """Carga datos de la hoja MANTRA del archivo REPORTE FTTH.xlsx
    Actualizado: 02/03/2026 - Ahora filtra por MES en lugar de FECHA"""
    excel_path = os.path.join(os.path.dirname(__file__), 'REPORTE FTTH.xlsx')
    
    try:
        df_mantra = pd.read_excel(excel_path, sheet_name='MANTRA')
        return df_mantra
    except Exception as e:
        return None

@st.cache_data(ttl=3600)    
def get_total_leads_and_conversion(mes_seleccionado="Noviembre"):
    """Obtiene total de leads y conversión para un mes específico"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 6589, 299  # Valores por defecto si no hay datos
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0, 0
    
    # Limpiar espacios en blanco
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    df_mes['NIVEL 3'] = df_mes['NIVEL 3'].astype(str).str.strip()
    
    # Total de leads para ese mes
    total_leads = len(df_mes)
    
    # Conversión: Con Cobertura + Contrato OK para ese mes
    df_conversion = df_mes[
        (df_mes['NIVEL 2'] == 'Con Cobertura') & 
        (df_mes['NIVEL 3'] == 'Contrato OK')
    ]
    total_conversion = len(df_conversion)
    
    return total_leads, total_conversion

@st.cache_data(ttl=3600)
def get_conversion_mantra_mes(mes_seleccionado="Noviembre"):
    """Calcula la conversión: Ventas Instaladas (DRIVE) / Con Cobertura (MANTRA)
    = Transacciones INSTALADAS en DRIVE / Registros con cobertura en MANTRA"""
    df_mantra = load_mantra_data()
    df_drive = load_drive_data()
    
    if df_mantra is None or df_mantra.empty or df_drive is None or df_drive.empty:
        return 0
    
    # Obtener Con Cobertura de MANTRA para el mes
    df_mes_mantra = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    df_mes_mantra['NIVEL 2'] = df_mes_mantra['NIVEL 2'].astype(str).str.strip()
    con_cobertura = len(df_mes_mantra[df_mes_mantra['NIVEL 2'] == 'Con Cobertura'])
    
    if con_cobertura == 0:
        return 0
    
    # Obtener Ventas INSTALADAS del DRIVE para el mes
    # Solo contar ESTADO = 'INSTALADO' (no PENDIENTE ni CANCELADO)
    ventas_instaladas = count_instaladas_con_regla(df_drive, None, mes_nombre=mes_seleccionado)
    
    if ventas_instaladas == 0:
        return 0
    
    # Conversión = Ventas Instaladas / Con Cobertura
    conversion_pct = round((ventas_instaladas / con_cobertura * 100)) if con_cobertura > 0 else 0
    return conversion_pct

@st.cache_data(ttl=3600)
def get_con_cobertura_count(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'Con Cobertura' para un mes específico"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Contar "Con Cobertura"
    con_cobertura = len(df_mes[df_mes['NIVEL 2'] == 'Con Cobertura'])
    
    return con_cobertura

@st.cache_data(ttl=3600)
def get_cancelados_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de cancelados para un mes específico usando columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    if mes_num is None:
        return 0
    
    # Filtrar por MES column si existe, sino por FECHA
    if 'MES' in df_drive.columns:
        df_mes = df_drive[
            (df_drive['MES'] == mes_seleccionado) &
            (df_drive['ESTADO'] == 'CANCELADO')
        ]
    else:
        # Fallback a FECHA
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        # Para Noviembre, incluir Octubre + Noviembre
        if mes_num == 11:
            df_mes = df_drive[
                ((df_drive['FECHA'].dt.month == 10) | (df_drive['FECHA'].dt.month == 11)) &
                (df_drive['ESTADO'] == 'CANCELADO')
            ]
        else:
            df_mes = df_drive[
                (df_drive['FECHA'].dt.month == mes_num) &
                (df_drive['ESTADO'] == 'CANCELADO')
            ]
    
    cancelados = len(df_mes)
    return cancelados

@st.cache_data
def get_instaladas_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de instaladas para un mes específico
    Regla: Solo INSTALADO (no incluye PENDIENTE)
    Filtra por columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    if mes_num is None:
        return 0
    
    # Para Noviembre, incluir Octubre + Noviembre
    es_noviembre = mes_num == 11
    instaladas = count_instaladas_con_regla(df_drive, mes_num, es_noviembre, mes_seleccionado)
    
    return instaladas

def get_ventas_generales_mes(mes_seleccionado="Noviembre"):
    """Obtiene el total de TODAS las transacciones del mes
    = INSTALADAS + PENDIENTES + CANCELADAS
    Filtra por columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Filtrar por mes usando columna MES
    if 'MES' in df_drive.columns:
        df_mes = df_drive[df_drive['MES'] == mes_seleccionado]
    else:
        # Fallback a FECHA si MES no existe
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_seleccionado, None)
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        df_mes = df_drive[df_drive['FECHA'].dt.month == mes_num]
    
    # Total de TODAS las transacciones
    total_transacciones = len(df_mes)
    return total_transacciones

@st.cache_data(ttl=3600)
def get_no_pago_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de NO PAGO para un mes específico usando columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    if mes_num is None:
        return 0
    
    # Limpiar espacios en blanco en MOTIVO CANCELACIÓN
    df_drive['MOTIVO CANCELACIÓN'] = df_drive['MOTIVO CANCELACIÓN'].astype(str).str.strip()
    
    # Filtrar por MES column si existe, sino por FECHA
    if 'MES' in df_drive.columns:
        df_mes = df_drive[
            (df_drive['MES'] == mes_seleccionado) &
            (df_drive['MOTIVO CANCELACIÓN'] == 'NO PAGO')
        ]
    else:
        # Fallback a FECHA
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        # Para Noviembre, incluir Octubre + Noviembre
        if mes_num == 11:
            df_mes = df_drive[
                ((df_drive['FECHA'].dt.month == 10) | (df_drive['FECHA'].dt.month == 11)) &
                (df_drive['MOTIVO CANCELACIÓN'] == 'NO PAGO')
            ]
        else:
            df_mes = df_drive[
                (df_drive['FECHA'].dt.month == mes_num) &
                (df_drive['MOTIVO CANCELACIÓN'] == 'NO PAGO')
            ]
    
    no_pago = len(df_mes)
    return no_pago

@st.cache_data(ttl=3600)
def get_no_responde_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'No Responde' para un mes específico desde MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco en NIVEL 1
    df_mes['NIVEL 1'] = df_mes['NIVEL 1'].astype(str).str.strip()
    
    # Contar "No Responde"
    no_responde = len(df_mes[df_mes['NIVEL 1'] == 'No Responde'])
    
    return no_responde

@st.cache_data(ttl=3600)
def get_no_especifica_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'No Especifica' para un mes específico desde MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco en NIVEL 2
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Contar "No Especifica"
    no_especifica = len(df_mes[df_mes['NIVEL 2'] == 'No Especifica'])
    
    return no_especifica

@st.cache_data(ttl=3600)
def get_sin_cobertura_mes(mes_seleccionado="Noviembre"):
    """Obtiene el conteo de 'Sin Cobertura' para un mes específico desde MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado]
    
    if df_mes.empty:
        return 0
    
    # Limpiar espacios en blanco en NIVEL 2
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    
    # Contar "Sin Cobertura"
    sin_cobertura = len(df_mes[df_mes['NIVEL 2'] == 'Sin Cobertura'])
    
    return sin_cobertura

@st.cache_data(ttl=3600)
def load_lista_metas():
    """Carga los datos de metas por mes de la hoja LISTA"""
    excel_path = os.path.join(os.path.dirname(__file__), 'REPORTE FTTH.xlsx')
    
    try:
        df_lista = pd.read_excel(excel_path, sheet_name='LISTA')
        return df_lista
    except Exception as e:
        return None

@st.cache_data(ttl=30)  # Reducir para asegurar datos frescos
def load_drive_data():
    """Carga datos de la hoja DRIVE del archivo REPORTE FTTH.xlsx"""
    excel_path = os.path.join(os.path.dirname(__file__), 'REPORTE FTTH.xlsx')
    
    try:
        df_drive = pd.read_excel(excel_path, sheet_name='DRIVE')
        return df_drive
    except Exception as e:
        return None

def count_instaladas_con_regla(df, fecha_mes_num, fecha_mes_es_noviembre=False, mes_nombre="Enero"):
    """
    Cuenta instaladas aplicando regla para todos los meses.
    
    Regla de VENTAS INSTALADAS DEL MES:
    - Solo INSTALADO
    - Sin considerar PENDIENTE
    - Filtra por columna MES (no por FECHA)
    
    Fórmula: COUNT(ESTADO='INSTALADO')
    
    Args:
        df: DataFrame del DRIVE
        fecha_mes_num: número del mes (deprecated, usa mes_nombre)
        fecha_mes_es_noviembre: si incluir Oct+Nov (deprecated)
        mes_nombre: nombre del mes ('Enero', 'Diciembre', etc)
    
    Returns:
        int: cantidad de instaladas según la regla
    """
    # Preparar dataframe
    df_temp = df.copy()
    df_temp['ESTADO'] = df_temp['ESTADO'].astype(str).str.strip()
    
    # Filtrar por columna MES (no por FECHA)
    if 'MES' in df_temp.columns:
        df_mes = df_temp[df_temp['MES'] == mes_nombre]
    else:
        # Fallback a filtro por FECHA si MES no existe
        df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
        if fecha_mes_es_noviembre:
            df_mes = df_temp[
                ((df_temp['FECHA'].dt.month == 10) | (df_temp['FECHA'].dt.month == 11))
            ]
        else:
            df_mes = df_temp[df_temp['FECHA'].dt.month == fecha_mes_num]
    
    # Aplicar regla: Solo INSTALADO
    df_instaladas = df_mes[df_mes['ESTADO'] == 'INSTALADO']
    
    return len(df_instaladas)

@st.cache_data(ttl=3600)
def get_meses_disponibles():
    """Obtiene lista de meses únicos disponibles en los datos con su año.
    Retorna lista de tuplas (mes_año, mes_nombre, año)"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return []
    
    df_temp = df_drive.copy()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # Extraer año y mes
    df_temp['AÑO'] = df_temp['FECHA'].dt.year
    df_temp['MES_NUM'] = df_temp['FECHA'].dt.month
    
    # Obtener combinaciones únicas de año y mes
    meses_unicos = df_temp.groupby(['AÑO', 'MES_NUM']).size().reset_index(name='count')
    meses_unicos = meses_unicos.sort_values(['AÑO', 'MES_NUM'], ascending=[False, False])
    
    # Mapear números a nombres
    mes_nombres = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
        5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
        9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    
    # Crear lista con formato "Mes Año"
    meses_disponibles = []
    for _, row in meses_unicos.iterrows():
        año = int(row['AÑO'])
        mes_num = int(row['MES_NUM'])
        mes_nombre = mes_nombres[mes_num]
        mes_año = f"{mes_nombre} {año}"
        meses_disponibles.append((mes_año, mes_nombre, año, mes_num))
    
    return meses_disponibles

@st.cache_data(ttl=3600)
def debug_instaladas_por_dia(mes_seleccionado="Febrero", dia=3):
    """Función de debug para ver qué registros hay en un día específico"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return None
    
    df_temp = df_drive.copy()
    
    # Sin limpiezas, solo conversion
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # Filtrar por mes y día
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    df_temp['FECHA_MES'] = df_temp['FECHA'].dt.month
    df_temp['FECHA_DIA'] = df_temp['FECHA'].dt.day
    
    # Filtrar por mes y día
    df_filtrado = df_temp[(df_temp['FECHA_MES'] == mes_num) & (df_temp['FECHA_DIA'] == dia)]
    
    # Retornar TODOS los registros sin filtrar
    return df_filtrado

@st.cache_data(ttl=60)  # Reducir a 60 segundos para debug
def get_instaladas_por_semana(mes_seleccionado="Noviembre"):
    """Obtiene VENTAS por DÍA para un mes específico.
    VENTAS = todos los registros del mes (sin importar PAGO o ESTADO)
    Retorna un DataFrame con día y cantidad de ventas
    Filtra por fecha actual para no mostrar registros futuros"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    # Conversión simple, sin limpieza excesiva
    df_temp = df_drive.copy()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # Mapeo de meses
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    
    mes_num = mes_numeros.get(mes_seleccionado, None)
    if mes_num is None:
        return pd.DataFrame()
    
    # Filtrar por fecha válida
    df_temp = df_temp[df_temp['FECHA'].notna()]
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Extraer mes y año de FECHA
    df_temp['FECHA_MES'] = df_temp['FECHA'].dt.month
    df_temp['FECHA_AÑO'] = df_temp['FECHA'].dt.year
    df_temp['FECHA_DIA'] = df_temp['FECHA'].dt.day
    
    # Filtrar por mes exacto
    df_mes = df_temp[df_temp['FECHA_MES'] == mes_num].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Si hay múltiples años, tomar el más reciente
    año_filtro = df_mes['FECHA_AÑO'].max()
    df_mes = df_mes[df_mes['FECHA_AÑO'] == año_filtro]
    
    # Filtrar VENTAS - todos los registros sin importar PAGO o ESTADO
    df_ventas = df_mes.copy()
    
    if df_ventas.empty:
        return pd.DataFrame()
    
    # Validar días válidos del mes
    if mes_num == 12:
        último_día_mes = pd.Timestamp(year=año_filtro+1, month=1, day=1) - pd.DateOffset(days=1)
    else:
        último_día_mes = pd.Timestamp(year=año_filtro, month=mes_num+1, day=1) - pd.DateOffset(days=1)
    
    último_día_válido = último_día_mes.day
    
    # Filtrar días válidos
    df_ventas = df_ventas[(df_ventas['FECHA_DIA'] >= 1) & (df_ventas['FECHA_DIA'] <= último_día_válido)]
    
    if df_ventas.empty:
        return pd.DataFrame()
    
    # Contar por día
    df_dias = df_ventas.groupby('FECHA_DIA').size().reset_index(name='INSTALADAS')
    df_dias.columns = ['DIA', 'INSTALADAS']
    
    # Crear etiquetas
    mes_nombres_cortos = {
        1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr',
        5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Ago',
        9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
    }
    mes_str = mes_nombres_cortos[mes_num]
    
    df_dias['DIA_ETIQUETA'] = df_dias['DIA'].astype(str) + ' ' + mes_str
    
    # Retornar ordenado por DIA (número), no por DIA_ETIQUETA (string)
    result = df_dias.sort_values('DIA')[['DIA_ETIQUETA', 'INSTALADAS']]
    result.columns = ['DIA', 'INSTALADAS']
    
    return result

@st.cache_data(ttl=3600)
def get_comparativo_semanas_multiples_meses():
    """Obtiene un comparativo de instaladas por DÍA para todos los meses disponibles.
    Retorna un DataFrame con día y cantidad por cada mes
    Filtra por fecha actual para no mostrar registros futuros"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    df_temp = df_drive.copy()
    df_temp['ESTADO'] = df_temp['ESTADO'].astype(str).str.strip()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Filtrar solo instaladas
    df_instaladas = df_temp[df_temp['ESTADO'] == 'INSTALADO'].copy()
    
    if df_instaladas.empty:
        return pd.DataFrame()
    
    # Obtener columna MES si existe, sino calcularla
    if 'MES' not in df_instaladas.columns:
        mes_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        df_instaladas['MES'] = df_instaladas['FECHA'].dt.month.map(mes_nombres)
    
    # Extraer día del mes
    df_instaladas['DIA'] = df_instaladas['FECHA'].dt.day
    
    # Agrupar por mes y día
    df_pivot = df_instaladas.groupby(['MES', 'DIA']).size().reset_index(name='INSTALADAS')
    
    # Crear tabla pivote (meses como columnas, días como filas)
    df_comparativo = df_pivot.pivot(index='DIA', columns='MES', values='INSTALADAS').fillna(0).astype(int)
    
    # Ordenar por día
    df_comparativo = df_comparativo.sort_index()
    
    return df_comparativo

def get_comparativo_acumulativo_multiples_meses():
    """Obtiene un comparativo ACUMULATIVO de instaladas para todos los meses disponibles.
    Retorna un DataFrame con día y cantidad acumulada por cada mes
    Filtra por fecha actual para no mostrar registros futuros"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return pd.DataFrame()
    
    df_temp = df_drive.copy()
    df_temp['ESTADO'] = df_temp['ESTADO'].astype(str).str.strip()
    df_temp['FECHA'] = pd.to_datetime(df_temp['FECHA'], errors='coerce')
    
    # FILTRO POR FECHA ACTUAL - no mostrar fechas futuras
    fecha_actual = pd.Timestamp.today()
    df_temp = df_temp[df_temp['FECHA'] <= fecha_actual]
    
    # Filtrar solo instaladas
    df_instaladas = df_temp[df_temp['ESTADO'] == 'INSTALADO'].copy()
    
    if df_instaladas.empty:
        return pd.DataFrame()
    
    # Obtener columna MES si existe, sino calcularla
    if 'MES' not in df_instaladas.columns:
        mes_nombres = {
            1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
            5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
            9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
        }
        df_instaladas['MES'] = df_instaladas['FECHA'].dt.month.map(mes_nombres)
    
    # Extraer día del mes
    df_instaladas['DIA'] = df_instaladas['FECHA'].dt.day
    
    # Agrupar por mes y día
    df_pivot = df_instaladas.groupby(['MES', 'DIA']).size().reset_index(name='INSTALADAS')
    
    # Crear tabla pivote (meses como columnas, días como filas)
    df_comparativo = df_pivot.pivot(index='DIA', columns='MES', values='INSTALADAS').fillna(0).astype(int)
    
    # Ordenar por día
    df_comparativo = df_comparativo.sort_index()
    
    # Calcular acumulados para cada mes
    df_acumulativo = df_comparativo.cumsum()
    
    return df_acumulativo

@st.cache_data(ttl=3600)
def get_nombres_alternativos(asesor):
    """Obtiene múltiples variantes del nombre del asesor para búsqueda flexible"""
    nombres = [asesor.strip()]
    # Agregar variante sin números al final (ej: ST2_VTP -> ST_VTP)
    import re
    nombre_sin_num = re.sub(r'(\d+)(_VTP)$', r'\2', asesor)
    if nombre_sin_num != asesor:
        nombres.append(nombre_sin_num)
    # Agregar variante con número (ej: ST_VTP -> ST2_VTP)
    nombre_con_num = re.sub(r'(_VTP)$', r'2_VTP', asesor.replace('2_VTP', '_VTP'))
    if nombre_con_num != asesor and '2_VTP' in nombre_con_num:
        nombres.append(nombre_con_num)
    return nombres

def get_pendientes_asesor_mes(asesor, mes_seleccionado="Enero"):
    """Obtiene cantidad de transacciones PENDIENTE por asesor para un mes"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_drive['ASESOR'] = df_drive['ASESOR'].astype(str).str.strip()
    asesor = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor)
    
    # Filtrar por mes y asesor (probando múltiples nombres)
    if 'MES' in df_drive.columns:
        df_mes_asesor = df_drive[(df_drive['MES'] == mes_seleccionado) & (df_drive['ASESOR'].isin(nombres_alternativos))]
    else:
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_seleccionado, None)
        if mes_num is None:
            return 0
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        df_mes_asesor = df_drive[(df_drive['FECHA'].dt.month == mes_num) & (df_drive['ASESOR'].isin(nombres_alternativos))]
    
    # Contar PENDIENTE
    df_mes_asesor['ESTADO'] = df_mes_asesor['ESTADO'].astype(str).str.strip()
    pendientes = len(df_mes_asesor[df_mes_asesor['ESTADO'] == 'PENDIENTE'])
    return pendientes

def get_leads_asesor_mes(asesor, mes_seleccionado="Enero"):
    """Obtiene el total de leads asignados a un asesor en un mes de MANTRA"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_mantra['Agente'] = df_mantra['Agente'].astype(str).str.strip()
    asesor = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor)
    
    # Obtener datos del asesor en MANTRA para el mes
    df_mes_asesor = df_mantra[(df_mantra['Mes'] == mes_seleccionado) & (df_mantra['Agente'].isin(nombres_alternativos))]
    
    # Total de leads (registros)
    total_leads = len(df_mes_asesor)
    return total_leads

def get_con_cobertura_asesor_mes(asesor, mes_seleccionado="Enero"):
    """Obtiene la cantidad de leads con cobertura para un asesor en un mes"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_mantra['Agente'] = df_mantra['Agente'].astype(str).str.strip()
    asesor = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor)
    
    # Obtener datos del asesor en MANTRA para el mes
    df_mes_asesor = df_mantra[(df_mantra['Mes'] == mes_seleccionado) & (df_mantra['Agente'].isin(nombres_alternativos))]
    
    if df_mes_asesor.empty:
        return 0
    
    # Limpiar espacios en NIVEL 2
    df_mes_asesor['NIVEL 2'] = df_mes_asesor['NIVEL 2'].astype(str).str.strip()
    
    # Contar "Con Cobertura"
    con_cobertura = len(df_mes_asesor[df_mes_asesor['NIVEL 2'] == 'Con Cobertura'])
    return con_cobertura

def get_conversion_asesor_mes(asesor, mes_seleccionado="Noviembre"):
    """Calcula la conversión por asesor: Contrato OK / Con Cobertura (de MANTRA)
    Usando datos de MANTRA únicamente"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return 0
    
    # Limpiar espacios en los nombres de asesor
    df_mantra['Agente'] = df_mantra['Agente'].astype(str).str.strip()
    asesor = asesor.strip()
    
    # Obtener nombres alternativos
    nombres_alternativos = get_nombres_alternativos(asesor)
    
    # Obtener datos del asesor en MANTRA para el mes
    df_mes_mantra = None
    for nombre in nombres_alternativos:
        df_temp = df_mantra[(df_mantra['Mes'] == mes_seleccionado) & (df_mantra['Agente'] == nombre)].copy()
        if not df_temp.empty:
            df_mes_mantra = df_temp
            break
    
    if df_mes_mantra is None or df_mes_mantra.empty:
        return 0
    
    # Limpiar niveles
    df_mes_mantra['NIVEL 2'] = df_mes_mantra['NIVEL 2'].astype(str).str.strip()
    df_mes_mantra['NIVEL 3'] = df_mes_mantra['NIVEL 3'].astype(str).str.strip()
    
    # Con Cobertura
    con_cobertura = len(df_mes_mantra[df_mes_mantra['NIVEL 2'] == 'Con Cobertura'])
    
    if con_cobertura == 0:
        return 0
    
    # Contrato OK (Con Cobertura + Contrato OK)
    contrato_ok = len(df_mes_mantra[
        (df_mes_mantra['NIVEL 2'] == 'Con Cobertura') & 
        (df_mes_mantra['NIVEL 3'] == 'Contrato OK')
    ])
    
    # Conversión = Contrato OK / Con Cobertura
    conversion_pct = round((contrato_ok / con_cobertura * 100)) if con_cobertura > 0 else 0
    return conversion_pct

@st.cache_data(ttl=3600)
def get_datos_mantra_mes(mes_seleccionado="Febrero"):
    """Obtiene datos detallados de MANTRA para un mes específico sin agregación"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Limpiar espacios en blanco
    df_mes['Agente'] = df_mes['Agente'].astype(str).str.strip()
    df_mes['NIVEL 1'] = df_mes['NIVEL 1'].astype(str).str.strip()
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    df_mes['NIVEL 3'] = df_mes['NIVEL 3'].astype(str).str.strip()
    
    return df_mes

@st.cache_data(ttl=3600)
def get_casos_por_agente_nivel(mes_seleccionado="Febrero"):
    """Obtiene casos por agente y nivel (1, 2, 3) desde MANTRA
    Retorna un DataFrame con información detallada para análisis"""
    df_mantra = load_mantra_data()
    
    if df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # Filtrar por mes
    df_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mes.empty:
        return pd.DataFrame()
    
    # Limpiar espacios en blanco
    df_mes['Agente'] = df_mes['Agente'].astype(str).str.strip()
    df_mes['NIVEL 1'] = df_mes['NIVEL 1'].astype(str).str.strip()
    df_mes['NIVEL 2'] = df_mes['NIVEL 2'].astype(str).str.strip()
    df_mes['NIVEL 3'] = df_mes['NIVEL 3'].astype(str).str.strip()
    
    # Agrupar por Agente
    agentes = df_mes['Agente'].unique()
    
    datos = []
    for agente in sorted(agentes):
        df_agente = df_mes[df_mes['Agente'] == agente]
        total_casos = len(df_agente)
        
        # Contar por NIVEL 1
        nivel1_counts = df_agente['NIVEL 1'].value_counts().to_dict()
        
        # Contar por NIVEL 2
        nivel2_counts = df_agente['NIVEL 2'].value_counts().to_dict()
        
        # Contar por NIVEL 3
        nivel3_counts = df_agente['NIVEL 3'].value_counts().to_dict()
        
        datos.append({
            'Agente': agente,
            'Total Casos': total_casos,
            'NIVEL 1': nivel1_counts,
            'NIVEL 2': nivel2_counts,
            'NIVEL 3': nivel3_counts
        })
    
    return pd.DataFrame(datos)

def calculate_drive_metrics(metas_dict, mes_filtro=None, mes_nombre=None):

    """
    Calcula Cumplimiento y Efectividad por asesor usando datos de DRIVE
    
    Cumplimiento = INSTALADAS / META
    Efectividad = INSTALADAS / (INSTALADAS + CANCELADAS)
    """
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return {}
    
    # Extraer columnas necesarias
    df_drive = df_drive[['FECHA', 'MES', 'ASESOR', 'ESTADO', 'PAGO']].copy()
    
    # Filtrar por mes usando columna MES si está disponible
    if mes_nombre and 'MES' in df_drive.columns:
        df_drive = df_drive[df_drive['MES'] == mes_nombre]
    elif mes_filtro:
        # Fallback a FECHA si MES no existe
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        df_drive = df_drive[df_drive['FECHA'].dt.month == mes_filtro]
    
    # Contar INSTALADOS por asesor (solo INSTALADO, sin PENDIENTE)
    df_drive_temp = df_drive.copy()
    df_drive_temp['ESTADO'] = df_drive_temp['ESTADO'].astype(str).str.strip()
    
    # Solo INSTALADO
    df_instalados = df_drive_temp[df_drive_temp['ESTADO'] == 'INSTALADO']
    instalados_por_asesor = df_instalados.groupby('ASESOR').size()
    
    # Contar CANCELADOS por asesor
    cancelados_por_asesor = df_drive[df_drive['ESTADO'] == 'CANCELADO'].groupby('ASESOR').size()
    
    # Calcular métricas
    metricas = {}
    for asesor, meta in metas_dict.items():
        instalados = instalados_por_asesor.get(asesor, 0)
        cancelados = cancelados_por_asesor.get(asesor, 0)
        
        # Cumplimiento = INSTALADAS / META
        cumplimiento = round((instalados / meta * 100) if meta > 0 else 0)
        
        # Efectividad = Nueva fórmula: Contrato OK / Con Cobertura (de MANTRA)
        efectividad = get_conversion_asesor_mes(asesor, mes_nombre)
        
        metricas[asesor] = {
            'instalados': instalados,
            'cancelados': cancelados,
            'cumplimiento': cumplimiento,
            'efectividad': efectividad
        }
    
    return metricas

# Cargar datos
def load_data(mes_seleccionado=None):
    # Cargar la hoja LISTA para obtener metas por mes
    df_lista = load_lista_metas()
    
    # Crear diccionario de metas para el mes seleccionado
    metas_dict = {}
    
    if df_lista is not None and not df_lista.empty:
        # Filtrar por el mes seleccionado
        df_mes_metas = df_lista[df_lista['Mes'] == mes_seleccionado].copy()
        
        # Limpiar espacios en blanco de Asesor y convertir Meta a numérico
        df_mes_metas['Asesor'] = df_mes_metas['Asesor'].astype(str).str.strip()
        df_mes_metas['Meta'] = pd.to_numeric(df_mes_metas['Meta'], errors='coerce').fillna(0)
        
        # Crear diccionario {Asesor: Meta} SOLO con los asesores activos en este mes
        for idx, row in df_mes_metas.iterrows():
            metas_dict[row['Asesor']] = int(row['Meta'])
    
    # Si no hay datos para el mes en LISTA, retornar vacío
    if not metas_dict:
        metas_dict = {}
    
    # Determinar número de mes
    mes_numeros = {
        'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
        'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
        'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
    }
    mes_num = mes_numeros.get(mes_seleccionado, None)
    
    # Obtener métricas de DRIVE filtrando por mes (ahora con mes_nombre)
    metricas = calculate_drive_metrics(metas_dict, mes_filtro=mes_num, mes_nombre=mes_seleccionado)
    
    # Construir DataFrame
    empleados = []
    metas = []
    cumplimientos = []
    efectividades = []
    instaladas = []
    canceladas = []
    
    for empleado, meta in metas_dict.items():
        empleados.append(empleado)
        metas.append(meta)
        
        if empleado in metricas:
            cumplimientos.append(metricas[empleado]['cumplimiento'])
            efectividades.append(metricas[empleado]['efectividad'])
            instaladas.append(metricas[empleado]['instalados'])
            canceladas.append(metricas[empleado]['cancelados'])
        else:
            cumplimientos.append(0)
            efectividades.append(0)
            instaladas.append(0)
            canceladas.append(0)
    
    data = {
        'Asesor': empleados,
        'Meta': metas,
        'Instaladas': instaladas,
        'Canceladas': canceladas,
        'Cumplimiento': cumplimientos,
        'Efectividad': efectividades
    }
    
    df = pd.DataFrame(data)
    df['% Meta Alcanzado'] = (df['Cumplimiento'] / 100 * 100).astype(int)
    df['Diferencia'] = df['Cumplimiento'] - 100
    return df

@st.cache_data(ttl=3600)
def load_data_codigo_carga(mes_seleccionado=None):
    """Carga datos agrupados por CODIGO DE CARGA (Agente) para un mes exacto.
    Incluye TODOS los agentes de MANTRA, incluso aquellos sin registros en DRIVE.
    - LEADS vienen de MANTRA (cantidad de registros por Agente)
    - VENTAS = todos los registros del mes (sin importar PAGO o ESTADO)
    - PEND = cantidad de PENDIENTES
    Filtra por columna MES exacto en ambas hojas."""
    df_drive = load_drive_data()
    df_mantra = load_mantra_data()
    
    if df_drive is None or df_drive.empty or df_mantra is None or df_mantra.empty:
        return pd.DataFrame()
    
    # ============= LEADS DESDE MANTRA =============
    # Filtrar por MES exacto en MANTRA
    df_mantra_mes = df_mantra[df_mantra['Mes'] == mes_seleccionado].copy()
    
    if df_mantra_mes.empty:
        return pd.DataFrame()
    
    # Limpiar espacios en blanco en Agente y estandarizar (agregar _VTP si no lo tiene)
    df_mantra_mes['Agente'] = df_mantra_mes['Agente'].astype(str).str.strip()
    # Estandarizar: si no termina con _VTP, agregar lo
    df_mantra_mes['Agente'] = df_mantra_mes['Agente'].apply(
        lambda x: x if x.endswith('_VTP') else x + '_VTP'
    )
    
    # Agrupar por Agente y contar LEADS
    leads_dict = df_mantra_mes.groupby('Agente').size().to_dict()
    
    # ============= ESTADÍSTICAS DESDE DRIVE =============
    # Filtrar por MES exacto en DRIVE
    df_drive_mes = df_drive[df_drive['MES'] == mes_seleccionado].copy()
    
    if df_drive_mes.empty:
        df_drive_mes = pd.DataFrame()
    else:
        # Limpiar espacios en blanco en columnas clave
        df_drive_mes['CODIGO DE CARGA'] = df_drive_mes['CODIGO DE CARGA'].astype(str).str.strip()
        df_drive_mes['ESTADO'] = df_drive_mes['ESTADO'].astype(str).str.strip()
        df_drive_mes['PAGO'] = df_drive_mes['PAGO'].astype(str).str.strip()
        df_drive_mes['FECHA'] = pd.to_datetime(df_drive_mes['FECHA'], errors='coerce')
        
        # Mapeo de meses a números para validar fecha
        mes_numeros_inv = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros_inv.get(mes_seleccionado, None)
        
        # Filtrar por fecha real del mes (excluir fechas que no pertenecen al mes)
        if mes_num:
            df_drive_mes = df_drive_mes[df_drive_mes['FECHA'].dt.month == mes_num].copy()
    
    # Agrupar por CODIGO DE CARGA y contar estados
    grupos = []
    
    # Obtener todos los agentes únicos de MANTRA (que son los CODIGO DE CARGA)
    agentes_unicos = sorted(leads_dict.keys())
    
    for agente in agentes_unicos:
        # Leads desde MANTRA
        leads = leads_dict.get(agente, 0)
        
        # Con cobertura desde MANTRA (NIVEL 2 = 'Con Cobertura')
        df_agente_mantra = df_mantra_mes[df_mantra_mes['Agente'] == agente]
        df_agente_mantra['NIVEL 2'] = df_agente_mantra['NIVEL 2'].astype(str).str.strip()
        con_cobertura = len(df_agente_mantra[df_agente_mantra['NIVEL 2'] == 'Con Cobertura'])
        
        # Inicializar contadores
        ventas = 0
        pendientes = 0
        
        # Si hay datos en DRIVE, buscar registros del agente
        if not df_drive_mes.empty:
            df_agente = df_drive_mes[df_drive_mes['CODIGO DE CARGA'] == agente]
            
            if not df_agente.empty:
                # VENTAS = todos los registros sin importar PAGO o ESTADO
                ventas = len(df_agente)
                
                # PENDIENTES = registros con ESTADO='PENDIENTE'
                pendientes = len(df_agente[df_agente['ESTADO'] == 'PENDIENTE'])
        
        grupos.append({
            'CODIGO_CARGA': agente,
            'LEADS': leads,
            'CON_COBERTURA': con_cobertura,
            'VENTAS': ventas,
            'PENDIENTES': pendientes
        })
    
    if not grupos:
        return pd.DataFrame()
    
    df_resultado = pd.DataFrame(grupos)
    
    # Calcular % Conversión de Ventas respecto a Leads: (VENTAS / LEADS) * 100
    df_resultado['CONV_VENTAS'] = (df_resultado['VENTAS'] / df_resultado['LEADS'] * 100).round(0).astype(int)
    
    # Calcular % Conversión de Ventas respecto a Con Cobertura: (VENTAS / CON_COBERTURA) * 100
    # Evitar división por cero
    df_resultado['CONV_VENTAS_COB'] = df_resultado.apply(
        lambda row: int((row['VENTAS'] / row['CON_COBERTURA'] * 100)) if row['CON_COBERTURA'] > 0 else 0,
        axis=1
    )
    
    # Calcular ventas necesarias para llegar al 10%: (LEADS * 0.10) - VENTAS
    df_resultado['VENTAS_FALTA_10'] = ((df_resultado['LEADS'] * 0.10) - df_resultado['VENTAS']).round(0).astype(int)
    # Si ya alcanzó el 10%, mostrar 0
    df_resultado['VENTAS_FALTA_10'] = df_resultado['VENTAS_FALTA_10'].apply(lambda x: max(0, x))
    
    # Ordenar por VENTAS de mayor a menor
    df_resultado = df_resultado.sort_values('VENTAS', ascending=False).reset_index(drop=True)
    
    # Agregar posición
    df_resultado.insert(0, 'POS', range(1, len(df_resultado) + 1))
    
    return df_resultado

# Estilos mejorados con tema moderno y premium
st.markdown("""
<style>
    :root {
        --primary-color: #0066cc;
        --secondary-color: #00d4ff;
        --success-color: #10b981;
        --warning-color: #f59e0b;
        --danger-color: #ef4444;
        --dark-bg: #0f172a;
        --light-bg: #f8fafc;
        --card-bg: #ffffff;
        --text-primary: #1e293b;
        --text-secondary: #64748b;
        --border-color: #e2e8f0;
    }

    * {
        font-family: 'Segoe UI', 'Helvetica Neue', sans-serif;
        margin: 0;
        padding: 0;
    }

    body {
        background-color: #f0f4f8;
    }

    /* HEADER PRINCIPAL */
    .header-container {
        background: linear-gradient(135deg, #0066cc 0%, #00d4ff 100%);
        padding: 40px 30px;
        border-radius: 15px;
        color: white;
        box-shadow: 0 10px 40px rgba(0, 102, 204, 0.2);
        margin-bottom: 30px;
        position: relative;
        overflow: hidden;
    }

    .header-container::before {
        content: '';
        position: absolute;
        top: -50%;
        right: -10%;
        width: 300px;
        height: 300px;
        background: rgba(255, 255, 255, 0.1);
        border-radius: 50%;
        z-index: 0;
    }

    .header-content {
        position: relative;
        z-index: 1;
    }

    .header-title {
        font-size: 2.8em;
        font-weight: 800;
        margin-bottom: 10px;
        letter-spacing: -0.5px;
        text-shadow: 0 2px 10px rgba(0, 0, 0, 0.1);
    }

    .header-subtitle {
        font-size: 1.3em;
        font-weight: 400;
        opacity: 0.95;
        margin-top: 5px;
    }

    /* KPI CARDS */
    .kpi-container {
        display: grid;
        grid-template-columns: repeat(6, 1fr);
        gap: 12px;
        margin-bottom: 30px;
    }

    .kpi-card {
        background: white;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.06);
        border: 1px solid #e2e8f0;
        transition: all 0.3s ease;
        position: relative;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        min-height: 160px;
    }

    .kpi-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 12px 24px rgba(0, 102, 204, 0.15);
        border-color: #0066cc;
    }

    .kpi-icon {
        font-size: 1.6em;
        margin-bottom: 8px;
    }

    .kpi-value {
        font-size: 2.2em;
        font-weight: 800;
        color: #0066cc;
        margin: 8px 0;
        font-family: 'Courier New', monospace;
        line-height: 1.2;
    }

    .kpi-label {
        font-size: 0.75em;
        color: #64748b;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        line-height: 1.3;
        word-wrap: break-word;
    }

    /* SECCIÓN TÍTULOS */
    .section-title {
        font-size: 1.6em;
        font-weight: 700;
        color: #1e293b;
        margin: 30px 0 20px 0;
        padding-bottom: 12px;
        border-bottom: 3px solid #0066cc;
        display: inline-block;
    }

    /* TABLA META */
    .meta-tabla {
        background: white;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        border: 1px solid #e2e8f0;
    }

    .meta-tabla table {
        width: 100%;
        border-collapse: collapse;
    }

    .meta-tabla thead {
        background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);
        color: white;
        font-weight: 700;
    }

    .meta-tabla th {
        padding: 8px 6px;
        text-align: left;
        font-size: 0.8em;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .meta-tabla td {
        padding: 5px 6px;
        border-bottom: 1px solid #f1f5f9;
        font-size: 0.8em;
    }

    .meta-tabla tbody tr {
        transition: background-color 0.2s ease;
    }

    .meta-tabla tbody tr:hover {
        background-color: #f8fafc;
    }

    .meta-tabla tbody tr:nth-child(odd) {
        background-color: #f9fafc;
    }

    .meta-valor {
        font-weight: 700;
        text-align: center;
        border-radius: 6px;
        padding: 2px 6px;
        display: inline-block;
        min-width: 32px;
        font-size: 0.8em;
        background: linear-gradient(135deg, #f59e0b 0%, #f97316 100%);
        color: white;
    }

    .meta-total {
        background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);
        color: white;
        font-weight: 800 !important;
    }

    .meta-total-row {
        background: linear-gradient(135deg, #f0f9ff 0%, #e0f2fe 100%);
        font-weight: 700;
    }

    /* CARDS DE GRÁFICOS */
    .chart-container {
        background: white;
        border-radius: 12px;
        padding: 20px;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        border: 1px solid #e2e8f0;
    }

    .chart-title {
        font-size: 1.2em;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 15px;
        display: flex;
        align-items: center;
        gap: 10px;
    }

    /* TABLA RESUMEN */
    .resumen-tabla {
        background: white;
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.08);
        border: 1px solid #e2e8f0;
    }

    .resumen-tabla table {
        width: 100%;
        border-collapse: collapse;
    }

    .resumen-tabla thead {
        background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);
        color: white;
    }

    .resumen-tabla th {
        padding: 10px 6px;
        text-align: center;
        font-size: 0.7em;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.3px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }

    .resumen-tabla td {
        padding: 8px 6px;
        border-bottom: 1px solid #f1f5f9;
        text-align: center;
        font-weight: 600;
        font-size: 0.85em;
    }

    .resumen-tabla tbody tr:hover {
        background-color: #f8fafc;
    }

    /* INDICADORES DE ESTADO */
    .status-excellent {
        color: #10b981;
        background: #dcfce7;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 700;
    }

    .status-good {
        color: #f59e0b;
        background: #fef3c7;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 700;
    }

    .status-poor {
        color: #ef4444;
        background: #fee2e2;
        padding: 4px 8px;
        border-radius: 6px;
        font-weight: 700;
    }

    /* BOTONES */
    .filter-button {
        padding: 10px 16px;
        border: 2px solid #e2e8f0;
        border-radius: 8px;
        background: white;
        color: #1e293b;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.2s ease;
    }

    .filter-button:hover {
        border-color: #0066cc;
        color: #0066cc;
    }

    /* DIVIDER */
    .section-divider {
        margin: 40px 0;
        border: none;
        border-top: 2px solid #e2e8f0;
    }

    /* FOOTER */
    .footer-container {
        text-align: center;
        color: #94a3b8;
        margin-top: 50px;
        padding: 30px;
        font-size: 0.9em;
    }

    .footer-container p {
        margin: 5px 0;
    }

    /* STREAMLIT CUSTOMIZATION */
    [data-testid="stDataFrame"] {
        border-radius: 12px !important;
        overflow: hidden !important;
    }

    /* SELECTBOX STYLING */
    .stSelectbox {
        border-radius: 8px !important;
    }

    /* TABS */
    [data-testid="stTabs"] {
        margin-top: 20px;
    }

    /* SCROLLBAR */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }

    ::-webkit-scrollbar-track {
        background: #f1f5f9;
    }

    ::-webkit-scrollbar-thumb {
        background: #cbd5e1;
        border-radius: 10px;
    }

    ::-webkit-scrollbar-thumb:hover {
        background: #94a3b8;
    }

    @media (max-width: 1400px) {
        .kpi-container {
            grid-template-columns: repeat(3, 1fr);
        }
    }

    @media (max-width: 768px) {
        .kpi-container {
            grid-template-columns: repeat(2, 1fr);
        }

        .header-title {
            font-size: 2em;
        }
    }
</style>
""", unsafe_allow_html=True)

# Filtros mejorados con layout dinámico
st.markdown("### ⚙️ Filtros y Opciones")
col_filtros = st.columns(3, gap="medium")

with col_filtros[0]:
    mes = st.selectbox("📅 Selecciona Mes", ["Noviembre", "Diciembre", "Enero", "Febrero", "Marzo"], index=4)

# Mapeo de meses a años
mes_año_map = {
    "Noviembre": "Noviembre 2025",
    "Diciembre": "Diciembre 2025",
    "Enero": "Enero 2026",
    "Febrero": "Febrero 2026",
    "Marzo": "Marzo 2026"
}

# Header mejorado - Dinámico
st.markdown(f"""
<div class="header-container">
    <div class="header-content">
        <div class="header-title">🌐 WORLDTEL</div>
        <div class="header-subtitle">Dashboard de Cumplimiento Mensual - {mes_año_map[mes]}</div>
    </div>
    <div style="position: absolute; right: 250px; top: 50%; transform: translateY(-50%); color: white; font-size: 3.8em; font-weight: 800; letter-spacing: -0.5px;">BITEL - FTTH</div>
</div>
""", unsafe_allow_html=True)

# Cargar datos con el mes seleccionado
df = load_data(mes)

with col_filtros[1]:
    opciones_asesores = ["Todos"] + sorted(df['Asesor'].unique())
    asesor_seleccionado = st.selectbox("👤 Filtrar por Asesor", opciones_asesores)

vista = "Completa"

# Obtener valores del Excel basado en el mes seleccionado
total_leads_excel, total_conversion_excel = get_total_leads_and_conversion(mes)

# KPI Cards mejorados - Datos del asesor seleccionado o totales
def get_cumplimiento_total_mes(mes_nombre):
    """Calcula el cumplimiento total del mes: (Total Instaladas / Total de Metas) * 100"""
    df_lista = load_lista_metas()
    df_drive = load_drive_data()
    
    if df_lista is None or df_lista.empty or df_drive is None or df_drive.empty:
        return 0
    
    try:
        # Mapear nombre del mes a número
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_nombre)
        
        if mes_num is None:
            return 0
        
        # Obtener total de metas para el mes
        df_mes_metas = df_lista[df_lista['Mes'] == mes_nombre]
        total_metas = df_mes_metas['Meta'].sum()
        
        if total_metas == 0:
            return 0
        
        # Aplicar regla: INSTALADO - pasando mes_nombre para filtrar por MES column
        es_noviembre = mes_num == 11
        total_instaladas = count_instaladas_con_regla(df_drive, mes_num, es_noviembre, mes_nombre)
        
        # Calcular cumplimiento
        cumplimiento_total = round((total_instaladas / total_metas * 100))
        
        return cumplimiento_total
    except Exception as e:
        return 0

def get_efectividad_mes(mes_nombre):
    """Calcula la efectividad para un mes: INSTALADAS/(INSTALADAS+CANCELADAS)
    Donde INSTALADAS = INSTALADO (sin PENDIENTE)"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    try:
        # Mapear nombre del mes a número
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_nombre)
        
        if mes_num is None:
            return 0
        
        # Filtrar por MES en lugar de FECHA
        if 'MES' in df_drive.columns:
            df_filtrado = df_drive[df_drive['MES'] == mes_nombre]
        else:
            # Fallback a FECHA si MES no existe
            df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
            # Para Noviembre, incluir Octubre + Noviembre
            if mes_num == 11:
                df_filtrado = df_drive[
                    ((df_drive['FECHA'].dt.month == 10) | (df_drive['FECHA'].dt.month == 11))
                ]
            else:
                # Para otros meses, solo ese mes
                df_filtrado = df_drive[
                    (df_drive['FECHA'].dt.month == mes_num)
                ]
        
        # Contar instaladas con regla (solo INSTALADO)
        instaladas = count_instaladas_con_regla(df_filtrado, mes_num, mes_num == 11, mes_nombre)
        canceladas = len(df_filtrado[df_filtrado['ESTADO'] == 'CANCELADO'])
        
        # Calcular efectividad
        total_transacciones = instaladas + canceladas
        if total_transacciones > 0:
            efectividad = round((instaladas / total_transacciones * 100))
        else:
            efectividad = 0
        
        return efectividad
    except Exception as e:
        return 0
        
        # Calcular efectividad
        total_transacciones = instaladas + canceladas
        if total_transacciones > 0:
            efectividad = round((instaladas / total_transacciones * 100))
        else:
            efectividad = 0
        
        return efectividad
    except Exception as e:
        return 0

def get_ventas_mes(mes_nombre):
    """Obtiene el total de instaladas para un mes específico del DRIVE
    Donde INSTALADAS = Solo INSTALADO (no incluye PENDIENTE)
    Filtra por columna MES"""
    df_drive = load_drive_data()
    
    if df_drive is None or df_drive.empty:
        return 0
    
    try:
        # Mapear nombre del mes a número
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes_nombre)
        
        if mes_num is None:
            return 0
        
        # Convertir FECHA a datetime
        df_drive['FECHA'] = pd.to_datetime(df_drive['FECHA'], errors='coerce')
        
        # Para Noviembre, sumar Octubre + Noviembre
        es_noviembre = mes_num == 11
        total = count_instaladas_con_regla(df_drive, mes_num, es_noviembre, mes_nombre)
        return total
    except Exception as e:
        return 0

st.markdown("")  # Espaciador

col1, col2, col3, col4, col5, col6, col7 = st.columns(7, gap="small")

if asesor_seleccionado == "Todos":
    # Preparar datos según vista
    df_vista = df[['Asesor', 'Cumplimiento']].copy()
    
    if vista == "Top 5":
        df_vista = df_vista.nlargest(5, 'Cumplimiento')
    elif vista == "Últimos 5":
        df_vista = df_vista.nsmallest(5, 'Cumplimiento')
    
    # Obtener asesores en la vista
    asesores_vista = df_vista['Asesor'].tolist()
    
    # Obtener ventas totales, efectividad y cumplimiento total del mes actual desde DRIVE
    # SIN filtrar por asesores - mostrar TOTALES de TODOS
    df_drive_filtrado = load_drive_data()
    
    if df_drive_filtrado is not None and not df_drive_filtrado.empty:
        # NO filtrar por asesores - calcular para TODOS
        # df_drive_filtrado = df_drive_filtrado[df_drive_filtrado['ASESOR'].isin(asesores_vista)]
        
        # Calcular métricas para esta vista
        df_drive_filtrado['FECHA'] = pd.to_datetime(df_drive_filtrado['FECHA'], errors='coerce')
        
        # Determinar número de mes
        mes_numeros = {
            'Enero': 1, 'Febrero': 2, 'Marzo': 3, 'Abril': 4,
            'Mayo': 5, 'Junio': 6, 'Julio': 7, 'Agosto': 8,
            'Septiembre': 9, 'Octubre': 10, 'Noviembre': 11, 'Diciembre': 12
        }
        mes_num = mes_numeros.get(mes, None)
        
        # Filtrar por columna MES (en lugar de por FECHA)
        if 'MES' in df_drive_filtrado.columns:
            df_mes_filtrado = df_drive_filtrado[df_drive_filtrado['MES'] == mes]
        else:
            # Fallback a FECHA si MES no existe
            df_drive_filtrado['FECHA'] = pd.to_datetime(df_drive_filtrado['FECHA'], errors='coerce')
            if mes_num == 11:
                df_mes_filtrado = df_drive_filtrado[
                    ((df_drive_filtrado['FECHA'].dt.month == 10) | (df_drive_filtrado['FECHA'].dt.month == 11))
                ]
            else:
                df_mes_filtrado = df_drive_filtrado[df_drive_filtrado['FECHA'].dt.month == mes_num]
        
        # Ventas (instaladas) - aplicando regla: Solo INSTALADO
        ventas_total = count_instaladas_con_regla(df_mes_filtrado, mes_num, mes_num == 11, mes)
        
        # Efectividad - Nueva fórmula: Contrato OK / Con Cobertura (de MANTRA)
        efectividad_mes = get_conversion_mantra_mes(mes)
        
        # Cumplimiento - calcular contra TODAS las metas del mes
        df_lista = load_lista_metas()
        df_mes_metas = df_lista[df_lista['Mes'] == mes]
        total_metas = df_mes_metas['Meta'].sum()
        cumplimiento_total = round((ventas_total / total_metas * 100)) if total_metas > 0 else 0
        
        # Ventas generales (total de todas las transacciones)
        ventas_generales = get_ventas_generales_mes(mes)
    else:
        ventas_total = 0
        efectividad_mes = 0
        cumplimiento_total = 0
        ventas_generales = 0
    
    kpis = [
        (f"{total_leads_excel:,}", "📋 Leads", col1),
        (str(get_con_cobertura_count(mes)), "🌐 Con Cobertura", col2),
        (f"{total_conversion_excel}", "✅ Ventas Del Mes", col3),
        (str(ventas_total), "💰 Ventas Instaladas Del Mes", col4),
        (str(ventas_generales), "📈 Ventas Generales Del Mes", col5),
        (f"{efectividad_mes}%", "⭐ Conversión de Ventas", col6),
        (f"{cumplimiento_total}%", "🎯 Cumplimiento", col7),
    ]
else:
    asesor_data = df[df['Asesor'] == asesor_seleccionado].iloc[0]
    cumpl_val = int(asesor_data['Cumplimiento'])
    efect_val = int(asesor_data['Efectividad'])
    instaladas_asesor = int(asesor_data['Instaladas'])
    leads_asesor = get_leads_asesor_mes(asesor_seleccionado, mes)
    con_cobertura_asesor = get_con_cobertura_asesor_mes(asesor_seleccionado, mes)
    
    kpis = [
        (str(leads_asesor), "📋 Total Leads", col1),
        (str(con_cobertura_asesor), "🌐 Con Cobertura", col2),
        (str(int(asesor_data['Meta'])), "🏆 Meta", col3),
        (f"{cumpl_val}%", "✅ Cumplimiento", col4),
        (f"{efect_val}%", "⭐ Conv. Ventas", col5),
        (str(instaladas_asesor), "💰 Instaladas", col6),
        ("🟢 Excelente" if cumpl_val >= 70 else "🟡 Bueno" if cumpl_val >= 50 else "🔴 Bajo", "📈 Estado", col7),
    ]

for valor, label, col in kpis:
    with col:
        st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-label">{label.split(' ', 1)[1]}</div>
            <div class="kpi-value">{valor}</div>
        </div>
        """, unsafe_allow_html=True)

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Sección principal con 3 columnas mejorada - RESPONSIVO
st.markdown("### 📑 Análisis de Desempeño por Agente")
col1, col2, col3 = st.columns([0.8, 1.6, 1.6], gap="medium")

# Columna 1: Meta Mensual
with col1:
    st.markdown('<div class="chart-title">📈 Meta Mensual</div>', unsafe_allow_html=True)
    tabla_meta = df[['Asesor', 'Meta']].copy()
    tabla_meta = tabla_meta.sort_values('Meta', ascending=False).reset_index(drop=True)
    tabla_meta.index = tabla_meta.index + 1
    
    # Crear HTML para la tabla personalizada
    html_tabla = '<div class="meta-tabla" style="width: auto; max-width: none;"><table><thead><tr><th>Pos</th><th>Asesor</th><th style="text-align: center;">Meta</th></tr></thead><tbody>'
    
    for idx, row in tabla_meta.iterrows():
        asesor = row['Asesor']
        meta = int(row['Meta'])
        html_tabla += f'<tr><td style="font-weight: 700; text-align: center; color: #0066cc;">#{idx}</td><td style="font-weight: 600;">{asesor}</td><td style="text-align: center;"><div class="meta-valor">{meta}</div></td></tr>'
    
    # Agregar fila de totales
    total_meta = int(tabla_meta['Meta'].sum())
    html_tabla += f'<tr style="background-color: #e0e7ff; font-weight: 700; border-top: 2px solid #0066cc;"><td style="text-align: center; color: #0066cc;">∑</td><td style="font-weight: 700; color: #0066cc;">TOTAL</td><td style="text-align: center; font-weight: 700; color: #0066cc;"><div class="meta-valor" style="background-color: #0066cc; color: white;">{total_meta}</div></td></tr>'
    
    html_tabla += '</tbody></table></div>'
    
    st.markdown(html_tabla, unsafe_allow_html=True)

# Columna 2: Cumplimiento por Agente
with col2:
    st.markdown('<div class="chart-title">🎯 Cumplimiento por Agente (%)</div>', unsafe_allow_html=True)
    df_sorted = df.sort_values('Cumplimiento', ascending=True)
    
    # Crear colores degradados basados en cumplimiento
    def get_color(val):
        if val >= 100:
            return '#10b981'  # Verde
        elif val >= 75:
            return '#f59e0b'  # Naranja
        elif val >= 50:
            return '#f97316'  # Naranja oscuro
        else:
            return '#ef4444'  # Rojo
    
    colors = [get_color(x) for x in df_sorted['Cumplimiento']]
    
    fig_cumpl = go.Figure()
    fig_cumpl.add_trace(go.Bar(
        y=df_sorted['Asesor'],
        x=df_sorted['Cumplimiento'],
        orientation='h',
        marker=dict(
            color=colors,
            line=dict(color='white', width=2),
            opacity=0.85
        ),
        text=df_sorted['Cumplimiento'].apply(lambda x: f'{x}%'),
        textposition='outside',
        textfont=dict(size=13, color='#1e293b', family='Arial', weight='bold'),
        hovertemplate='<b>%{y}</b><br><b>Cumplimiento:</b> <b>%{x}%</b><extra></extra>',
        name=''
    ))
    fig_cumpl.update_layout(
        height=580,
        margin=dict(l=160, r=50, t=20, b=20),
        showlegend=False,
        xaxis_title="",
        xaxis=dict(
            gridcolor='rgba(0,0,0,0.05)',
            showgrid=True,
            zeroline=False,
            tickfont=dict(size=11, color='#64748b'),
            range=[0, 130],
            ticksuffix='%'
        ),
        yaxis=dict(
            tickfont=dict(size=9, color='#1e293b'),
            automargin=True
        ),
        plot_bgcolor='rgba(248, 250, 252, 0.5)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=11, family='Arial', color='#1e293b'),
        hovermode='closest'
    )
    st.plotly_chart(fig_cumpl, use_container_width=True, config={'displayModeBar': False})

# Columna 3: Conversión de Ventas por Agente
with col3:
    st.markdown('<div class="chart-title">⭐ Conversión de Ventas por Agente (%)</div>', unsafe_allow_html=True)
    df_sorted_eff = df.sort_values('Efectividad', ascending=True)
    
    colors_eff = [get_color(x) for x in df_sorted_eff['Efectividad']]
    
    fig_eff = go.Figure()
    fig_eff.add_trace(go.Bar(
        y=df_sorted_eff['Asesor'],
        x=df_sorted_eff['Efectividad'],
        orientation='h',
        marker=dict(
            color=colors_eff,
            line=dict(color='white', width=2),
            opacity=0.85
        ),
        text=df_sorted_eff['Efectividad'].apply(lambda x: f'{x}%'),
        textposition='outside',
        textfont=dict(size=13, color='#1e293b', family='Arial', weight='bold'),
        hovertemplate='<b>%{y}</b><br><b>Efectividad:</b> <b>%{x}%</b><extra></extra>',
        name=''
    ))
    fig_eff.update_layout(
        height=580,
        margin=dict(l=50, r=50, t=20, b=20),
        showlegend=False,
        xaxis_title="",
        xaxis=dict(
            gridcolor='rgba(0,0,0,0.05)',
            showgrid=True,
            zeroline=False,
            tickfont=dict(size=11, color='#64748b'),
            range=[0, 130],
            ticksuffix='%'
        ),
        yaxis=dict(
            tickfont=dict(size=9, color='#1e293b'),
            automargin=True
        ),
        plot_bgcolor='rgba(248, 250, 252, 0.5)',
        paper_bgcolor='rgba(0,0,0,0)',
        font=dict(size=11, family='Arial', color='#1e293b'),
        hovermode='closest'
    )
    st.plotly_chart(fig_eff, use_container_width=True, config={'displayModeBar': False})

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ============= ANÁLISIS DE INSTALADAS POR SEMANA =============
st.markdown("### 📊 Análisis de Instaladas por Semana")

# Crear dos tabs: uno para mes individual y otro para comparativo
tab1, tab2 = st.tabs(["Análisis por Semana (Mes)", "Comparativo Multi-Mes"])

# TAB 1: Análisis de semanas para un mes seleccionado
with tab1:
    # Obtener meses disponibles
    meses_disp = get_meses_disponibles()
    
    if meses_disp:
        # Crear lista de opciones con formato "Mes Año"
        opciones_meses = [mes_año for mes_año, _, _, _ in meses_disp]
        
        col_mes_sel, col_espacio = st.columns([2, 3])
        with col_mes_sel:
            mes_seleccionado_display = st.selectbox(
                "Selecciona un mes para analizar:",
                opciones_meses,
                index=0,  # Primer mes disponible por defecto
                key="mes_analisis"
            )
        
        # Encontrar el mes_nombre del mes seleccionado
        mes_nombre_analisis = next((mes_nombre for mes_año, mes_nombre, _, _ in meses_disp if mes_año == mes_seleccionado_display), None)
    else:
        st.warning("No hay datos disponibles en los registros")
        mes_nombre_analisis = None
    
    if mes_nombre_analisis:
        # Obtener datos de instaladas por día
        df_semanas = get_instaladas_por_semana(mes_nombre_analisis)
    
    if not df_semanas.empty and len(df_semanas) > 0:
        # Crear gráfico de barras
        fig_semanas = go.Figure()
        
        fig_semanas.add_trace(go.Bar(
            x=df_semanas['DIA'],
            y=df_semanas['INSTALADAS'],
            marker=dict(
                color=df_semanas['INSTALADAS'],
                colorscale='Blues',
                line=dict(color='white', width=2),
                opacity=0.85,
                showscale=True,
                colorbar=dict(
                    title="Ventas",
                    tickfont=dict(size=10),
                    thickness=15,
                    len=0.7
                )
            ),
            text=df_semanas['INSTALADAS'],
            textposition='outside',
            textfont=dict(size=13, color='#1e293b', family='Arial', weight='bold'),
            hovertemplate='<b>%{x}</b><br><b>Ventas:</b> <b>%{y}</b><extra></extra>',
            name='Ventas'
        ))
        
        fig_semanas.update_layout(
            title=dict(
                text=f"Distribución Diaria de Ventas - {mes_seleccionado_display}",
                font=dict(size=16, color='#1e293b', family='Arial'),
                x=0.5,
                xanchor='center'
            ),
            height=550,
            margin=dict(l=50, r=50, t=80, b=150),
            xaxis_title="Día",
            yaxis_title="Cantidad de Ventas",
            xaxis=dict(
                tickfont=dict(size=9, color='#64748b'),
                tickangle=-90,
            ),
            yaxis=dict(
                gridcolor='rgba(0,0,0,0.05)',
                showgrid=True,
                zeroline=False,
                tickfont=dict(size=11, color='#64748b'),
            ),
            plot_bgcolor='rgba(248, 250, 252, 0.5)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=11, family='Arial', color='#1e293b'),
            hovermode='x unified',
            showlegend=False
        )
        
        st.plotly_chart(fig_semanas, use_container_width=True, config={'displayModeBar': False})
        
        # Mostrar tabla de datos
        st.markdown("#### Detalle Diario")
        df_tabla_semanas = df_semanas.copy()
        df_tabla_semanas.columns = ['Día', 'Ventas']
        
        # Calcular estadísticas
        max_dia = df_tabla_semanas.loc[df_tabla_semanas['Ventas'].idxmax(), 'Día']
        max_valor = df_tabla_semanas['Ventas'].max()
        min_valor = df_tabla_semanas['Ventas'].min()
        promedio = df_tabla_semanas['Ventas'].mean()
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Día con + Ventas", max_dia, f"+{max_valor}")
        with col2:
            st.metric("Máximo por Día", max_valor)
        with col3:
            st.metric("Mínimo por Día", min_valor)
        with col4:
            st.metric("Promedio por Día", f"{promedio:.1f}")
        
        st.dataframe(
            df_tabla_semanas,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Día": st.column_config.TextColumn(width=200),
                "Instaladas": st.column_config.NumberColumn(width=150)
            }
        )
        
        # ============= COMPARATIVA POR HORARIO: FULL TIME vs PART TIME dentro del tab1 =============
        st.markdown('<div style="margin-top: 40px;"></div>', unsafe_allow_html=True)
        st.markdown("#### 📊 Comparativa por Horario: FULL TIME vs PART TIME")
        
        # Función para generar datos detallados por horario
        def generar_tabla_horario(mes_sel, meta_minima=60):
            df_lista = load_lista_metas()
            df_drive = load_drive_data()
            
            if df_lista is None or df_drive is None:
                return None, None
            
            # Preparar LISTA - limpiar espacios en Asesor y convertir Meta a numérico
            df_lista_clean = df_lista.copy()
            df_lista_clean['Asesor'] = df_lista_clean['Asesor'].astype(str).str.strip()
            df_lista_clean['Meta'] = pd.to_numeric(df_lista_clean['Meta'], errors='coerce').fillna(0)
            
            # Preparar DRIVE
            df_drive_clean = df_drive.copy()
            df_drive_clean['ASESOR'] = df_drive_clean['ASESOR'].astype(str).str.strip()
            df_drive_clean['ESTADO'] = df_drive_clean['ESTADO'].astype(str).str.strip()
            df_mes_drive = df_drive_clean[df_drive_clean['MES'] == mes_sel]
            
            # Obtener datos de LISTA
            df_mes_lista = df_lista_clean[df_lista_clean['Mes'] == mes_sel]
            
            # Clasificar asesores por horario: FULL TIME son meta 60 o meta 45, resto es PART TIME
            full_time = df_mes_lista[(df_mes_lista['Meta'] == 60) | (df_mes_lista['Meta'] == 45)]['Asesor'].tolist()
            part_time = df_mes_lista[(df_mes_lista['Meta'] != 60) & (df_mes_lista['Meta'] != 45)]['Asesor'].tolist()
            
            def procesar_horario(lista_asesoras):
                datos_tabla = []
                total_meta = 0
                total_instalado = 0
                total_pendiente = 0
                
                for idx, asesor in enumerate(lista_asesoras, 1):
                    # Meta
                    meta = df_mes_lista[df_mes_lista['Asesor'] == asesor]['Meta'].sum()
                    if meta == 0:
                        meta = 0
                    
                    # Instalado y Pendiente
                    df_asesor = df_mes_drive[df_mes_drive['ASESOR'] == asesor]
                    instalado = len(df_asesor[df_asesor['ESTADO'] == 'INSTALADO'])
                    pendiente = len(df_asesor[df_asesor['ESTADO'] == 'PENDIENTE'])
                    
                    # Alcance (Cumplimiento)
                    alcance = round((instalado / meta * 100)) if meta > 0 else 0
                    
                    datos_tabla.append({
                        'pos': idx,
                        'asesor': asesor,
                        'meta': int(meta),
                        'instalado': int(instalado),
                        'pendiente': int(pendiente),
                        'alcance': int(alcance)
                    })
                    
                    total_meta += meta
                    total_instalado += instalado
                    total_pendiente += pendiente
                
                # Ordenar datos por alcance de mayor a menor
                datos_tabla_ordenado = sorted(datos_tabla, key=lambda x: x['alcance'], reverse=True)
                
                # Actualizar posiciones después del ordenamiento
                for idx, item in enumerate(datos_tabla_ordenado, 1):
                    item['pos'] = idx
                
                return {
                    'datos': datos_tabla_ordenado,
                    'totales': {
                        'meta': int(total_meta),
                        'instalado': int(total_instalado),
                        'pendiente': int(total_pendiente),
                        'alcance': round((total_instalado / total_meta * 100)) if total_meta > 0 else 0
                    }
                }
            
            datos_full_time = procesar_horario(full_time)
            datos_part_time = procesar_horario(part_time)
            
            return datos_full_time, datos_part_time
        
        # Generar datos de ambos horarios
        datos_full_time, datos_part_time = generar_tabla_horario(mes_nombre_analisis)
        
        # Crear tablas HTML detalladas
        def crear_tabla_html_horario(datos_horario, nombre_horario, color_header, color_accent):
            if datos_horario is None or not datos_horario['datos']:
                return ""
            
            html = f'''<div style="margin: 20px 0; background: white; border-radius: 8px; overflow: auto; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
            <table style="width: 100%; border-collapse: collapse; font-family: Arial, sans-serif; min-width: 100%;">
            <thead>
                <tr style="background: {color_header}; color: white;">
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">Nº ASESOR</th>
                    <th style="padding: 12px; text-align: left; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2); min-width: 170px;">ASESOR</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">OBJETIVO</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">INSTALADO</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">PENDIENTE</th>
                    <th style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">% ALCANCE</th>
                </tr>
            </thead>
            <tbody>
            '''
            
            # Agregar filas de datos
            for item in datos_horario['datos']:
                color_fila = '#f9fafb' if item['pos'] % 2 == 0 else '#ffffff'
                alcance_color = '#10b981' if item['alcance'] >= 70 else '#f59e0b' if item['alcance'] >= 50 else '#ef4444'
                
                html += f'''<tr style="background-color: {color_fila}; border-bottom: 1px solid #e5e7eb;">
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; color: {color_accent};">#{item['pos']}</td>
                    <td style="padding: 10px 12px; text-align: left; font-weight: 500; font-size: 11px;">{item['asesor']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px;">{item['meta']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; color: #10b981;">{item['instalado']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; color: #f59e0b;">{item['pendiente']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-weight: 600; font-size: 11px; background-color: {alcance_color}22; color: {alcance_color}; border-radius: 4px;">{item['alcance']}%</td>
                </tr>'''
            
            # Agregar fila de totales
            totales = datos_horario['totales']
            alcance_total_color = '#10b981' if totales['alcance'] >= 70 else '#f59e0b' if totales['alcance'] >= 50 else '#ef4444'
            html += f'''<tr style="background: {color_header}; color: white; font-weight: 700;">
                    <td colspan="2" style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['meta']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['meta']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['instalado']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white;">{totales['pendiente']}</td>
                    <td style="padding: 10px 12px; text-align: center; font-size: 12px; color: white; background-color: {alcance_total_color}40; border-radius: 4px;">{totales['alcance']}%</td>
                </tr>
            </tbody>
            </table>
            </div>'''
            
            return html
        
        # Generar HTML para ambos horarios
        html_full_time = crear_tabla_html_horario(datos_full_time, "FULL TIME", "#3b82f6", "#3b82f6")
        html_part_time = crear_tabla_html_horario(datos_part_time, "PART TIME", "#8b5cf6", "#8b5cf6")
        
        # Mostrar tablas en dos columnas
        col_full_time, col_part_time = st.columns(2)
        
        with col_full_time:
            st.markdown('<h4 style="text-align: center; color: #3b82f6; margin-bottom: 10px;">⏰ FULL TIME </h4>', unsafe_allow_html=True)
            if html_full_time:
                st.markdown(html_full_time, unsafe_allow_html=True)
            else:
                st.info("No hay asesores FULL TIME")
        
        with col_part_time:
            st.markdown('<h4 style="text-align: center; color: #8b5cf6; margin-bottom: 10px;">⏰ PART TIME </h4>', unsafe_allow_html=True)
            if html_part_time:
                st.markdown(html_part_time, unsafe_allow_html=True)
            else:
                st.info("No hay asesores PART TIME")
    else:
        st.warning(f"No hay datos de instaladas para {mes_seleccionado_display}")

# TAB 2: Comparativo acumulativo entre meses
with tab2:
    st.markdown("#### 📈 Comparativa Acumulativa de Instaladas (Todos los Meses)")
    
    df_comparativo = get_comparativo_acumulativo_multiples_meses()
    
    if not df_comparativo.empty:
        # Crear gráfico de líneas para comparar meses
        fig_comparativo = go.Figure()
        
        # Agregar línea por cada mes
        for mes_col in df_comparativo.columns:
            fig_comparativo.add_trace(go.Scatter(
                x=df_comparativo.index,
                y=df_comparativo[mes_col],
                mode='lines+markers',
                name=mes_col,
                line=dict(width=2.5),
                marker=dict(size=6),
                hovertemplate='<b>Día %{x}</b><br><b>' + mes_col + ':</b> %{y} acumuladas<extra></extra>'
            ))
        
        fig_comparativo.update_layout(
            title=dict(
                text="Comparativa Acumulativa de Instaladas por Día (Todos los Meses)",
                font=dict(size=16, color='#1e293b', family='Arial'),
                x=0.5,
                xanchor='center'
            ),
            height=550,
            margin=dict(l=60, r=60, t=80, b=80),
            xaxis_title="Día del Mes",
            yaxis_title="Total Acumulado de Instaladas",
            xaxis=dict(
                tickfont=dict(size=11, color='#64748b'),
                tickmode='linear',
                tick0=1,
                dtick=2
            ),
            yaxis=dict(
                gridcolor='rgba(0,0,0,0.05)',
                showgrid=True,
                zeroline=False,
                tickfont=dict(size=11, color='#64748b'),
            ),
            plot_bgcolor='rgba(248, 250, 252, 0.5)',
            paper_bgcolor='rgba(0,0,0,0)',
            font=dict(size=11, family='Arial', color='#1e293b'),
            hovermode='x unified',
            legend=dict(
                x=1.02,
                y=1,
                xanchor='left',
                yanchor='top',
                bgcolor='rgba(255, 255, 255, 0.8)',
                bordercolor='#e2e8f0',
                borderwidth=1
            )
        )
        
        st.plotly_chart(fig_comparativo, use_container_width=True, config={'displayModeBar': False})
    else:
        st.warning("No hay datos suficientes para el comparativo de meses")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Tabla de detalle de asesores
st.markdown("### 👥 Detalle Completo de Asesores")

# Filtro para ordenamiento
criterio_orden = st.selectbox(
    "Ordenar por:",
    ["Cumplimiento (Mayor a Menor)", "Conversión (Mayor a Menor)"],
    key="criterio_orden"
)

df_detail = df[['Asesor', 'Meta', 'Instaladas', 'Canceladas', 'Cumplimiento', 'Efectividad']].copy()
df_detail['Cumpl%'] = df_detail['Cumplimiento'].astype(str) + '%'
df_detail['Efect%'] = df_detail['Efectividad'].astype(str) + '%'

# Agregar columna de Pendientes solo para el mes actual
# Obtener mes actual
meses_nombres = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
                 7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
mes_actual = meses_nombres[datetime.now().month]

# Solo agregar columna de Pendientes si el mes seleccionado es el mes actual
if mes == mes_actual:
    pendientes_list = []
    for asesor in df_detail['Asesor']:
        pendientes = get_pendientes_asesor_mes(asesor, mes)
        pendientes_list.append(pendientes)
    df_detail['Pendientes'] = pendientes_list

# Agregar columnas de Leads y Con Cobertura
leads_list = []
con_cobertura_list = []
for asesor in df_detail['Asesor']:
    leads = get_leads_asesor_mes(asesor, mes)
    con_cobertura = get_con_cobertura_asesor_mes(asesor, mes)
    leads_list.append(leads)
    con_cobertura_list.append(con_cobertura)
df_detail['Leads'] = leads_list
df_detail['Con Cobertura'] = con_cobertura_list

# Separar en Full Time (meta >= 55) y Part Time (meta < 55)
# Excepción: CARLACA, ISABEL y LAURA son FULL TIME aunque tengan meta 45
asesoras_fulltime_especial = ['ZIM_CARLACA_VTP', 'ZIM_ISABELPF_VTP', 'ZIM_LAURAVS_VTP']
condicion_fulltime = (df_detail['Meta'] >= 55) | (df_detail['Asesor'].isin(asesoras_fulltime_especial))
df_fulltime = df_detail[condicion_fulltime].copy()
df_parttime = df_detail[~condicion_fulltime].copy()

# Ordenar por el criterio seleccionado
if criterio_orden == "Conversión (Mayor a Menor)":
    df_fulltime = df_fulltime.sort_values('Efectividad', ascending=False).reset_index(drop=True)
    df_parttime = df_parttime.sort_values('Efectividad', ascending=False).reset_index(drop=True)
else:
    df_fulltime = df_fulltime.sort_values('Cumplimiento', ascending=False).reset_index(drop=True)
    df_parttime = df_parttime.sort_values('Cumplimiento', ascending=False).reset_index(drop=True)

# Función para generar tabla HTML
def generar_tabla_detalle(df_tabla, tipo_empleado):
    # Obtener mes actual
    meses_nombres = {1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
                     7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'}
    mes_actual = meses_nombres[datetime.now().month]
    
    # Verificar si mostrar columna de Pendientes (solo si es el mes actual)
    mostrar_pendientes = mes == mes_actual
    
    if mostrar_pendientes:
        html_tabla = '<div class="meta-tabla"><table><thead><tr><th style="width: 4%;">Pos</th><th style="width: 18%;">Asesor</th><th style="width: 6%;">Leads</th><th style="width: 8%;">Cob</th><th style="width: 6%;">Meta</th><th style="width: 7%;">Inst</th><th style="width: 7%;">Canc</th><th style="width: 7%;">Pend</th><th style="width: 8%;">Cumpl%</th><th style="width: 9%;">Conv%</th><th style="width: 12%;">Estado</th></tr></thead><tbody>'
    else:
        html_tabla = '<div class="meta-tabla"><table><thead><tr><th style="width: 5%;">Pos</th><th style="width: 22%;">Asesor</th><th style="width: 7%;">Leads</th><th style="width: 8%;">Cob</th><th style="width: 7%;">Meta</th><th style="width: 8%;">Inst</th><th style="width: 8%;">Canc</th><th style="width: 9%;">Cumpl%</th><th style="width: 10%;">Conv%</th><th style="width: 10%;">Estado</th></tr></thead><tbody>'

    for idx, (_, row) in enumerate(df_tabla.iterrows(), 1):
        asesor = row['Asesor']
        leads = int(row.get('Leads', 0))
        con_cobertura = int(row.get('Con Cobertura', 0))
        meta = int(row['Meta'])
        instaladas = int(row['Instaladas'])
        canceladas = int(row['Canceladas'])
        pendientes = int(row.get('Pendientes', 0)) if mostrar_pendientes else 0
        cumpl = int(row['Cumplimiento'])
        efect = int(row['Efectividad'])
        
        # Determinar estado
        if cumpl >= 70:
            estado = '<span class="status-excellent">✓ Excelente</span>'
            fila_bg = 'background-color: #f0fdf4;'
        elif cumpl >= 50:
            estado = '<span class="status-good">~ Bueno</span>'
            fila_bg = 'background-color: #fffbeb;'
        else:
            estado = '<span class="status-poor">✗ Bajo</span>'
            fila_bg = 'background-color: #fef2f2;'
        
        if mostrar_pendientes:
            html_tabla += f'''<tr style="{fila_bg}">
                <td style="font-weight: 700; text-align: center; color: #0066cc;">#{idx}</td>
                <td style="font-weight: 600;">{asesor}</td>
                <td style="text-align: center; font-weight: 600; color: #0066cc;">{leads}</td>
                <td style="text-align: center; font-weight: 600; color: #8b5cf6;">{con_cobertura}</td>
                <td style="text-align: center; font-weight: 600;">{meta}</td>
                <td style="text-align: center; font-weight: 600; color: #10b981;">{instaladas}</td>
                <td style="text-align: center; font-weight: 600; color: #ef4444;">{canceladas}</td>
                <td style="text-align: center; font-weight: 600; color: #f59e0b;">{pendientes}</td>
                <td style="text-align: center;"><div class="meta-valor">{cumpl}%</div></td>
                <td style="text-align: center;"><div class="meta-valor" style="background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);">{efect}%</div></td>
                <td style="text-align: center;">{estado}</td>
            </tr>'''
        else:
            html_tabla += f'''<tr style="{fila_bg}">
                <td style="font-weight: 700; text-align: center; color: #0066cc;">#{idx}</td>
                <td style="font-weight: 600;">{asesor}</td>
                <td style="text-align: center; font-weight: 600; color: #0066cc;">{leads}</td>
                <td style="text-align: center; font-weight: 600; color: #8b5cf6;">{con_cobertura}</td>
                <td style="text-align: center; font-weight: 600;">{meta}</td>
                <td style="text-align: center; font-weight: 600; color: #10b981;">{instaladas}</td>
                <td style="text-align: center; font-weight: 600; color: #ef4444;">{canceladas}</td>
                <td style="text-align: center;"><div class="meta-valor">{cumpl}%</div></td>
                <td style="text-align: center;"><div class="meta-valor" style="background: linear-gradient(135deg, #0066cc 0%, #0052a3 100%);">{efect}%</div></td>
                <td style="text-align: center;">{estado}</td>
            </tr>'''

    html_tabla += '</tbody></table></div>'
    return html_tabla

# Mostrar tabla Full Time
if not df_fulltime.empty:
    st.markdown("#### 💼 Asesores Full Time (8 horas - Meta ≥ 55)")
    html_fulltime = generar_tabla_detalle(df_fulltime, "Full Time")
    st.markdown(html_fulltime, unsafe_allow_html=True)
    st.markdown('<div style="margin: 15px 0;"></div>', unsafe_allow_html=True)

# Mostrar tabla Part Time
if not df_parttime.empty:
    st.markdown("#### 👨‍💼 Asesores Part Time (4 horas - Meta < 55)")
    html_parttime = generar_tabla_detalle(df_parttime, "Part Time")
    st.markdown(html_parttime, unsafe_allow_html=True)

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Mostrar tabla combinada de todos los asesores por CODIGO DE CARGA
st.markdown("#### 👥 Detalle Completo de Todos los Asesores")

# Cargar datos agrupados por CODIGO DE CARGA para el mes seleccionado
df_codigos_carga = load_data_codigo_carga(mes)

if not df_codigos_carga.empty:
    # Crear tabla HTML con el formato deseado
    def generar_tabla_codigos_carga(df_datos):
        """Genera tabla HTML para datos agrupados por CODIGO DE CARGA"""
        html = '''<div style="margin: 20px 0; background: white; border-radius: 8px; overflow: auto; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
        <table style="width: 100%; border-collapse: collapse; font-family: Arial, sans-serif;">
        <thead>
            <tr style="background: linear-gradient(135deg, #0066cc 0%, #00d4ff 100%); color: white;">
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">POS</th>
                <th style="padding: 14px; text-align: left; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2); min-width: 180px;">CODIGO CARGA</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">TOTAL DE LEADS</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);"># CON COBERTURA</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">%CONV. VENTAS</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px; border-right: 1px solid rgba(255,255,255,0.2);">TOTAL DE VENTAS</th>
                <th style="padding: 14px; text-align: center; font-weight: 700; font-size: 12px;">%CONV. VENTAS FINAL</th>
            </tr>
        </thead>
        <tbody>
        '''
        
        for idx, row in df_datos.iterrows():
            color_fila = '#f9fafb' if idx % 2 == 0 else '#ffffff'
            pos = int(row['POS'])
            codigo = row['CODIGO_CARGA']
            leads = int(row['LEADS'])
            con_cobertura = int(row['CON_COBERTURA'])
            ventas = int(row['VENTAS'])
            conv_ventas = int(row['CONV_VENTAS'])
            conv_ventas_cob = int(row['CONV_VENTAS_COB'])
            
            # Determinar color para ventas
            if ventas > 0:
                color_ventas = '#10b981'  # Verde
            else:
                color_ventas = '#64748b'  # Gris
            
            # Determinar color para conversión respecto a leads
            if conv_ventas >= 10:
                color_conv = '#10b981'  # Verde
            elif conv_ventas == 9:
                color_conv = '#f59e0b'  # Naranja
            else:
                color_conv = '#ef4444'  # Rojo
            
            # Determinar color para conversión respecto a con cobertura
            if conv_ventas_cob >= 10:
                color_conv_cob = '#10b981'  # Verde
            elif conv_ventas_cob == 9:
                color_conv_cob = '#f59e0b'  # Naranja
            else:
                color_conv_cob = '#ef4444'  # Rojo
            
            html += f'''<tr style="background-color: {color_fila}; border-bottom: 1px solid #e5e7eb;">
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; color: #0066cc;">#{pos}</td>
                <td style="padding: 12px; text-align: left; font-weight: 500; font-size: 12px;">{codigo}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px;">{leads}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px;">{con_cobertura}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; background-color: {color_conv_cob}22; color: {color_conv_cob}; border-radius: 4px;">{conv_ventas_cob}%</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; color: {color_ventas};">{ventas}</td>
                <td style="padding: 12px; text-align: center; font-weight: 600; font-size: 12px; background-color: {color_conv}22; color: {color_conv}; border-radius: 4px;">{conv_ventas}%</td>
            </tr>'''
        
        # Calcular y agregar fila de TOTALES
        total_leads = df_datos['LEADS'].sum()
        total_con_cobertura = df_datos['CON_COBERTURA'].sum()
        total_ventas = df_datos['VENTAS'].sum()
        total_conv_ventas_cob = int((total_ventas / total_con_cobertura * 100)) if total_con_cobertura > 0 else 0
        total_conv_ventas = int((total_ventas / total_leads * 100)) if total_leads > 0 else 0
        
        # Determinar color para conversión total respecto a con cobertura
        if total_conv_ventas_cob >= 10:
            color_conv_total_cob = '#10b981'  # Verde
        elif total_conv_ventas_cob == 9:
            color_conv_total_cob = '#f59e0b'  # Naranja
        else:
            color_conv_total_cob = '#ef4444'  # Rojo
        
        # Determinar color para conversión total respecto a leads
        if total_conv_ventas >= 10:
            color_conv_total = '#10b981'  # Verde
        elif total_conv_ventas == 9:
            color_conv_total = '#f59e0b'  # Naranja
        else:
            color_conv_total = '#ef4444'  # Rojo
        
        html += f'''<tr style="background: linear-gradient(135deg, #0066cc 0%, #00d4ff 100%); color: white; font-weight: 700;">
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;"></td>
            <td style="padding: 12px; text-align: left; font-weight: 700; font-size: 12px;">TOTAL</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">{total_leads}</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">{total_con_cobertura}</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; background-color: {color_conv_total_cob}40; color: white; border-radius: 4px;">{total_conv_ventas_cob}%</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px;">{total_ventas}</td>
            <td style="padding: 12px; text-align: center; font-weight: 700; font-size: 12px; background-color: {color_conv_total}40; color: white; border-radius: 4px;">{total_conv_ventas}%</td>
        </tr>'''
        
        html += '''</tbody>
        </table>
        </div>'''
        
        return html
    
    # Generar y mostrar tabla
    html_tabla = generar_tabla_codigos_carga(df_codigos_carga)
    st.markdown(html_tabla, unsafe_allow_html=True)
    
    # Mostrar estadísticas generales
    st.markdown("#### 📊 Resumen General por Mes")
    col1, col2, col3, col4, col5 = st.columns(5)
    
    total_codigos = len(df_codigos_carga)
    total_leads = df_codigos_carga['LEADS'].sum()
    total_ventas = df_codigos_carga['VENTAS'].sum()
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Códigos de Carga", total_codigos)
    with col2:
        st.metric("Total Leads", total_leads)
    with col3:
        st.metric("Total Ventas", total_ventas)
else:
    st.info("No hay datos disponibles para el mes seleccionado")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# ============= NUEVA SECCIÓN: ANÁLISIS DE CASOS POR NIVEL =============
st.markdown("### 📋 Análisis de Casos por Nivel (MANTRA)")

# Obtener datos detallados de MANTRA
df_mantra_mes = get_datos_mantra_mes(mes)

if not df_mantra_mes.empty:
    # Crear filtros múltiples en 4 columnas
    col1, col2, col3, col4 = st.columns(4, gap="small")
    
    with col1:
        agentes_unique = sorted(df_mantra_mes['Agente'].unique())
        agente_filtro = st.selectbox(
            "Agente",
            ["Todos"] + list(agentes_unique),
            key="agente_filtro_casos"
        )
    
    # Filtrar por agente para obtener valores únicos de niveles
    if agente_filtro == "Todos":
        df_temp = df_mantra_mes.copy()
    else:
        df_temp = df_mantra_mes[df_mantra_mes['Agente'] == agente_filtro]
    
    with col2:
        nivel1_unique = sorted(df_temp['NIVEL 1'].unique())
        nivel1_filtro = st.selectbox(
            "Nivel 1",
            ["Todos"] + list(nivel1_unique),
            key="nivel1_filtro_casos"
        )
    
    # Filtrar por nivel 1
    if nivel1_filtro == "Todos":
        df_temp2 = df_temp.copy()
    else:
        df_temp2 = df_temp[df_temp['NIVEL 1'] == nivel1_filtro]
    
    with col3:
        nivel2_unique = sorted(df_temp2['NIVEL 2'].unique())
        nivel2_filtro = st.selectbox(
            "Nivel 2",
            ["Todos"] + list(nivel2_unique),
            key="nivel2_filtro_casos"
        )
    
    # Filtrar por nivel 2
    if nivel2_filtro == "Todos":
        df_temp3 = df_temp2.copy()
    else:
        df_temp3 = df_temp2[df_temp2['NIVEL 2'] == nivel2_filtro]
    
    with col4:
        nivel3_unique = sorted(df_temp3['NIVEL 3'].unique())
        nivel3_filtro = st.multiselect(
            "Nivel 3",
            list(nivel3_unique),
            default=list(nivel3_unique),
            key="nivel3_filtro_casos"
        )
    
    # Aplicar todos los filtros
    df_filtrado = df_mantra_mes.copy()
    
    if agente_filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado['Agente'] == agente_filtro]
    
    if nivel1_filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado['NIVEL 1'] == nivel1_filtro]
    
    if nivel2_filtro != "Todos":
        df_filtrado = df_filtrado[df_filtrado['NIVEL 2'] == nivel2_filtro]
    
    if nivel3_filtro:
        df_filtrado = df_filtrado[df_filtrado['NIVEL 3'].isin(nivel3_filtro)]
    
    # Mostrar total de casos filtrados
    total_casos_filtrados = len(df_filtrado)
    st.markdown(f"### Total de Casos: **{total_casos_filtrados}**")
    
    # Mostrar vista previa (primeros 10 registros)
    if total_casos_filtrados > 0:
        col_preview, col_download = st.columns([4, 1])
        
        with col_preview:
            st.markdown("#### Vista Previa (Primeros 10 registros)")
            
            # Seleccionar columnas para mostrar
            cols_mostrar = ['Agente', 'NIVEL 1', 'NIVEL 2', 'NIVEL 3']
            if 'Telefono' in df_filtrado.columns:
                cols_mostrar.insert(1, 'Telefono')
            if 'Numero Caso' in df_filtrado.columns:
                cols_mostrar.insert(0, 'Numero Caso')
            
            df_preview = df_filtrado[cols_mostrar].head(10)
            st.dataframe(df_preview, use_container_width=True, hide_index=True)
            
            if total_casos_filtrados > 10:
                st.caption(f"Mostrando 10 de {total_casos_filtrados} registros")
        
        with col_download:
            st.markdown("#### Descargar")
            
            # Crear archivo Excel con todos los datos filtrados
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_filtrado.to_excel(writer, sheet_name='Casos Filtrados', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Casos Filtrados']
                
                # Estilos
                header_fill = PatternFill(start_color="0066cc", end_color="0066cc", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                
                # Aplicar estilos al encabezado
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                # Aplicar estilos a datos
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
                # Ajustar anchos de columna
                for col_num, col in enumerate(worksheet.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(col_num)
                    worksheet.column_dimensions[col_letter].width = 20
            
            buffer.seek(0)
            
            # Preparar nombre del archivo con agente y fecha
            nombre_asesor = agente_filtro if agente_filtro != "Todos" else "Todos"
            fecha_hoy = datetime.now().strftime('%d_%m_%Y')
            nombre_archivo = f"Casos_Filtrados_{nombre_asesor}_{mes}_{fecha_hoy}.xlsx"
            
            st.download_button(
                label="📥 Descargar Excel",
                data=buffer,
                file_name=nombre_archivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    else:
        st.warning("No hay casos que coincidan con los filtros seleccionados")

else:
    st.warning(f"No hay datos de casos disponibles para {mes}")

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Tabla de resumen mensual con expanders
st.markdown("### ⭐ Resumen Mensual Completo")

st.markdown('<div style="margin: 20px 0;"></div>', unsafe_allow_html=True)

# Obtener datos para cada mes
meses_disponibles = ['Noviembre', 'Diciembre', 'Enero', 'Febrero']
datos_meses = []
totales = {'Leads': 0, 'Contr': 0, 'Cober': 0}

for mes_nombre in meses_disponibles:
    leads, conversion = get_total_leads_and_conversion(mes_nombre)
    con_cobertura = get_con_cobertura_count(mes_nombre)
    cancelados = get_cancelados_mes(mes_nombre)
    instaladas = get_instaladas_mes(mes_nombre)
    no_pago = get_no_pago_mes(mes_nombre)
    no_responde = get_no_responde_mes(mes_nombre)
    no_especifica = get_no_especifica_mes(mes_nombre)
    sin_cobertura = get_sin_cobertura_mes(mes_nombre)
    datos_meses.append({
        'Mes': mes_nombre,
        'Leads': leads,
        'Cober': con_cobertura,
        'Contr': conversion,
        'Cancel': cancelados,
        'Pago': instaladas,
        'NoPago': no_pago,
        'NoResp': no_responde,
        'NoEsp': no_especifica,
        'SinCob': sin_cobertura
    })
    totales['Leads'] += leads
    totales['Cober'] += con_cobertura
    totales['Contr'] += conversion

# Construir tabla HTML de resumen sin expanders (solo datos de resumen)
html_resumen = '''
<div class="resumen-tabla">
<table>
<thead><tr>
<th>Mes</th>
<th>Leads</th>
<th>Cober</th>
<th>%Cob</th>
<th>Contr</th>
<th>%Conv</th>
<th>Cancel</th>
<th>Pagó</th>
<th>NoPag</th>
<th>Efect</th>
<th>NoResp</th>
<th>%NR</th>
<th>NoEsp</th>
<th>%NE</th>
<th>%SC</th>
</tr></thead><tbody>
'''

for dato in datos_meses:
    mes_nombre = dato['Mes']
    leads = dato['Leads']
    cober = dato['Cober']
    cob_pct = int(cober/leads*100) if leads > 0 else 0
    contr = dato['Contr']
    conv_pct = get_conversion_mantra_mes(mes_nombre)
    cancel = dato['Cancel']
    pago = dato['Pago']
    nopago = dato['NoPago']
    efect_pct = int(pago/(cancel+pago+nopago)*100) if (cancel+pago+nopago) > 0 else 0
    noresp = dato['NoResp']
    noresp_pct = int(noresp/leads*100) if leads > 0 else 0
    noesp = dato['NoEsp']
    noesp_pct = int(noesp/leads*100) if leads > 0 else 0
    sincob = dato['SinCob']
    sincob_pct = int(sincob/leads*100) if leads > 0 else 0
    
    html_resumen += f'''<tr>
    <td><strong>{mes_nombre}</strong></td>
    <td>{leads}</td>
    <td>{cober}</td>
    <td>{cob_pct}%</td>
    <td>{contr}</td>
    <td style="color: #0066cc; font-weight: 700;">{conv_pct}%</td>
    <td>{cancel}</td>
    <td>{pago}</td>
    <td>{nopago}</td>
    <td style="color: #0066cc; font-weight: 700;">{efect_pct}%</td>
    <td>{noresp}</td>
    <td>{noresp_pct}%</td>
    <td>{noesp}</td>
    <td>{noesp_pct}%</td>
    <td>{sincob_pct}%</td>
    </tr>'''

html_resumen += '</tbody></table></div>'

st.markdown(html_resumen, unsafe_allow_html=True)

st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

# Footer elegante
st.markdown("""
<div class="footer-container">
    <p><strong>Dashboard WORLD TEL</strong> - Sistema de Control de Cumplimiento Mensual</p>
    <p>📅 Periodo: Noviembre 2025 | 🕐 Actualizado: {}  | 👥 Total Empleados: 14</p>
    <p style="margin-top: 15px; opacity: 0.7;">© 2025 WORLD TEL | Todos los derechos reservados</p>
</div>
""".format(datetime.now().strftime("%d/%m/%Y %H:%M:%S")), unsafe_allow_html=True)
